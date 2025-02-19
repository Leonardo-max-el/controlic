from django.shortcuts import render,redirect
from django.http import HttpResponse
from django.contrib.auth.forms import UserCreationForm,AuthenticationForm
from django.contrib.auth.models import User
from .models import Inspeccionambientes
from django.http import JsonResponse
import openpyxl
from django.core.paginator import Paginator

from django.conf import settings
from openpyxl.utils import get_column_letter
from collections import defaultdict
import os

from datetime import datetime
from openpyxl.worksheet.copier import WorksheetCopy


# Imortamos el login para poder autenticar al usuario dentro del navegador
from django.contrib.auth import login, logout,authenticate
from django.db import IntegrityError


from flask import Flask

app = Flask(__name__)

@app.route('/')
def home():
    return "¡Hola desde la red WiFi!"

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=8000)  # '0.0.0.0' permite conexiones desde otros dispositivos




# Create your views here.
def home(request):
    return render(request, 'home.html')


def signup(request):

    if request.method == 'GET':
        return render (request, 'signup.html', {'form': UserCreationForm()})
    
    else:
        if request.POST['password1'] == request.POST['password2']:
            try:
                user = User.objects.create_user(username=request.POST['username'], password=request.POST['password1'])
                user.save()
                login(request, user)
                return redirect('registrar_inspeccion')
                # return render(request, 'signup.html', {'form': UserCreationForm(), 'error': 'Usuario creado correctamente'})
            except IntegrityError:
                return render(request, 'signup.html', {'form': UserCreationForm(), 'error': 'El usuario ya existe'})
        
        return render(request, 'signup.html', {'form': UserCreationForm(), 'error': 'Las contraseñas no coinciden'})
        

def clouses(request):
    logout(request)
    return redirect('home')

def login_view(request):
    if request.method == 'GET':
        return render(request, 'login_view.html',{
            'form':AuthenticationForm
    })
    else: 
        user = authenticate(
            request, username=request.POST['username'], 
            password=request.POST['password'])
        
        if user is None:
            return render(request, 'login_view.html',{
                'form':AuthenticationForm,
                'error': 'Usuario o contraseña Incorrecto'
            })
    
        else:
            login(request, user)
            return redirect('registrar_inspeccion')

# cambiar el laoratorio por ambientes

def registrar_inspeccion(request):
    ambientes = ["LABORATORIO DE CÓMPUTO", "AULA", "TALLER"]
    
    elementos = {
    "Equipo de Cómputo": "equipo_computo",
    "Proyector Multimedia": "proyector_multimedia",
    "Red": "red",
    "Fluido Eléctrico": "fluido_electrico",
    "Orden y Limpieza": "orden_limpieza",
    "Módulos": "modulos"
}



    if request.method == "POST":
        # Obtener los datos del formulario
        usuario = request.user if request.user.is_authenticated else None
        tipo_ambiente = request.POST.get("tipo_ambiente")
        ambiente = request.POST.get("ambiente")
        equipo_computo = request.POST.get("equipo_computo")
        proyector_multimedia = request.POST.get("proyector_multimedia")
        red = request.POST.get("red")
        fluido_electrico = request.POST.get("fluido_electrico")
        orden_limpieza = request.POST.get("orden_limpieza")
        modulos = request.POST.get("modulos")
        
        # Asegurar que observación no sea None ni esté vacía
        observacion = request.POST.get("observacion", "").strip()
        if not observacion:
            observacion = "NO SE ENCONTRARON OBSERVACIONES"

        # Guardar en la base de datos
        Inspeccionambientes.objects.create(
            usuario=usuario,
            tipo_ambiente=tipo_ambiente,
            ambiente=ambiente,
            equipo_computo=equipo_computo,
            proyector_multimedia=proyector_multimedia,
            red=red,
            fluido_electrico=fluido_electrico,
            orden_limpieza=orden_limpieza,
            modulos=modulos,
            observacion=observacion
        )

        return JsonResponse({"mensaje": "Registro guardado exitosamente"}, status=200)

    return render(request, "registrar_inspeccion.html", {"ambientes": ambientes, "elementos": elementos})




def read_inspeccion(request):
    inspecciones_list = Inspeccionambientes.objects.all()
    paginator = Paginator(inspecciones_list, 5)  # 10 elementos por página
    page_number = request.GET.get('page')
    inspecciones = paginator.get_page(page_number)
    
    return render(request, "read_inspeccion.html", {"inspecciones": inspecciones})



def exportar_inspeccion_con_plantilla(request):
    fecha_inicio = request.GET.get("fecha_inicio")
    fecha_fin = request.GET.get("fecha_fin")
    
    if not fecha_inicio or not fecha_fin:
        return HttpResponse("Debe proporcionar un rango de fechas.", status=400)
    
    # Ruta de la plantilla
    template_path = os.path.join(os.path.dirname(__file__), "media", "plantilla.xlsx")
    wb = openpyxl.load_workbook(template_path)
    
    # Filtrar inspecciones por rango de fechas
    inspecciones = Inspeccionambientes.objects.filter(fecha_registro__range=[fecha_inicio, fecha_fin])
    
    # Agrupar inspecciones por fecha
    inspecciones_por_fecha = defaultdict(list)
    for inspeccion in inspecciones:
        fecha_str = inspeccion.fecha_registro.strftime("%Y-%m-%d")
        inspecciones_por_fecha[fecha_str].append(inspeccion)
    
    for fecha, registros in inspecciones_por_fecha.items():
        # Crear una nueva hoja por cada fecha
        if fecha in wb.sheetnames:
            ws = wb[fecha]  # Si ya existe, usarla
        else:
            ws = wb.copy_worksheet(wb.active)  # Copiar la estructura de la plantilla
            ws.title = fecha  # Renombrar la hoja con la fecha
        
        fila = 11  # Suponiendo que los registros inician en la fila 11
        for inspeccion in registros:
            columnas_fijas = ["D", "E", "R", "S"]
            valores_fijos = [
                inspeccion.tipo_ambiente,
                inspeccion.ambiente,
                inspeccion.usuario.username,  
                inspeccion.observacion
            ]
            
            for col, valor in zip(columnas_fijas, valores_fijos):
                ws[f"{col}{fila}"] = valor
            
            valores_dinamicos = [
                inspeccion.equipo_computo, 
                inspeccion.proyector_multimedia, 
                inspeccion.red, 
                inspeccion.fluido_electrico, 
                inspeccion.orden_limpieza, 
                inspeccion.modulos
            ]
            
            for col_c, col_o, valor in zip(["F", "H", "J", "L", "N", "P"], ["G", "I", "K","M","O","Q"], valores_dinamicos):
                if valor == "C":
                    ws[f"{col_c}{fila}"] = "X"
                elif valor == "O":
                    ws[f"{col_o}{fila}"] = "X"
            
            fila += 1  # Pasar a la siguiente fila
    
    # Configurar la respuesta HTTP para la descarga
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="reporte_inspecciones_{fecha_inicio}_a_{fecha_fin}.xlsx"'
    
    wb.save(response)
    return response